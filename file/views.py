from django.shortcuts import render
from django.http import FileResponse
from cmdb.models import Project, Server, Service
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from django.utils.encoding import escape_uri_path
import unicodedata

# Create your views here.

def download_excel(request):
    file = open('static/file/template.xlsx', 'rb')
    response = FileResponse(file)
    response['Content-Type'] = 'application/octet-stream'
    response['Content-Disposition'] = 'attachment;filename="template.xlsx"'
    return response

#导入
def import_data(filename):
    print(filename)
    book = load_workbook('tmp/import/' + filename)
    sheet = book.active
    info_list = []
    count = 0
    for C1,C2,C3,C4,C5,C6,C7,C8,C9,C10,C11,C12,C13,C14,C15,C16,C17,C18,C19,C20,C21,C22,C23 in sheet[sheet.dimensions]:
        print(C1.value,C2.value,C3.value,C4.value,C5.value,C6.value,C7.value,C8.value,C9.value,C10.value,C11.value,C12.value,C13.value,C14.value,C15.value,C16.value,C17.value,C18.value,C19.value,C20.value,C21.value,C22.value,C23.value)
        count += 1
        print(count)
        if count >= 3:
            if C1.value is None or C2.value is None or C3.value is None or C12.value is None or C13.value is None or C23.value is None:
                info_list.append('第{}行,必填项存在空白,不导入该行.'.format(count))
            else:
                print('test')
                envs = {'生产环境':'prod', '联调环境':'uat', '开发环境':'dev', '测试环境':'test'}
                os_types = {'linux': 'l', 'windows': 'w'}
                mx_version = str(C3.value) + '.0'
                try:
                    proj, _ = Project.objects.get_or_create(name=C1.value,env=envs[C2.value],mx_version=mx_version,cp_addr=C4.value,
                                                  cp_admin=C5.value, cp_password=C6.value, monitor_addr=C7.value,
                                                  monitor_admin=C8.value, monitor_password=C9.value,ops_admin=C10.value,
                                                  ops_password=C11.value
                                                            )
                    server, _ = Server.objects.get_or_create(ip=C12.value,project=proj,os_type=os_types[C13.value],os=C14.value,cpu=C15.value,
                                                 mem=C16.value,disk=C17.value,remote_port=C18.value,root=C19.value,passwd=C20.value,
                                                 mac=C21.value,remark=C22.value
                                                 )
                    #service, _ = Service.objects.get_or_create(name=C23.value,server=server,service_version=C24.value,db_root=C25.value,db_passwd=C26.value,
                    #                              charcater=charcaters[C27.value],repl_name=C28.value,remark=C29.value,protocol_port=C30.value,
                    #                              )
                    for service in C23.value.split('/'):
                        service, _ = Service.objects.get_or_create(name=service, server=server)
                except:
                    info_list.append('导入到第{}行时出现未知问题,中断操作.'.format(count))
    info_list.append('共导入{}行,成功'.format(count-2))

    context = {
        'info_list': info_list,
    }
    return context

#上传文件
def import_excel(request):
    if request.method == 'GET':
        return render(request, 'file/upload.html', context={})
    elif request.method == 'POST':
        try:
            file = request.FILES['file']
            filename = file.name
            if filename.split('.')[1] != 'xlsx':
                context = {'error': '仅支持xlsx后缀文件'}
            else:
                with open('tmp/import/' + filename, 'wb') as f:
                    for chunk in file.chunks():
                        f.write(chunk)

                context = import_data(filename)
        except:
            context = {'error': '请选择文件'}

        return render(request, 'file/upload.html', context=context)

#导出
def export_excel(request):
    book = load_workbook('static/file/template_export.xlsx')
    sheet1 = book.get_sheet_by_name('服务器及应用信息')
    sheet2 = book.get_sheet_by_name('项目信息')
    project_id = request.GET['project_id']
    project = Project.objects.get(pk=project_id)
    servers = Server.objects.filter(project=project_id)

    server_start_count = 3
    service_start_count = 3
    for server in servers:
        services = Service.objects.filter(server=server.pk)
        server_inc_count = 0
        #当服务器无应用时,写入应用名写入None,不然导出时会出错
        if not services:
            row1 = (
                (server.ip, server.get_os_type_display(), server.os, server.cpu, server.mem, server.disk, server.remote_port, server.root, server.passwd, server.mac, server.remark,
                  'None', service.service_version, service.db_root, service.db_passwd, service.get_charcater_display(), service.repl_name, service.remark, protocol.upper(), port)
                )
            sheet1.append(row1)
            continue
        for service in services:
            service_inc_count = 0

            str_proto_port = unicodedata.normalize('NFKC', str(service.protocol_port))
            for protocol_port in str_proto_port.split('/'):
                try:
                    protocol, port = protocol_port.split('-')
                except:
                    protocol, port = ['', '']
                server_inc_count += 1
                service_inc_count += 1
                row1 = (
                    (server.ip, server.get_os_type_display(), server.os, server.cpu, server.mem, server.disk, server.remote_port, server.root, server.passwd, server.mac, server.remark,
                     service.name, service.service_version, service.db_root, service.db_passwd, service.get_charcater_display(), service.repl_name, service.remark, protocol.upper(), port)
                )
                sheet1.append(row1)
            service_end_count = service_start_count + service_inc_count - 1

            #"应用"合并并居中
            column_count = 11
            for i in list(map(chr, range(ord('L'), ord('R') + 1))):
                column_count += 1
                start_cell = str(i) + str(service_start_count)
                end_cell = str(i) + str(service_end_count)
                sheet1.merge_cells(start_cell + ':' + end_cell)
                cell_name = str(i) + str(service_start_count)
                cell = sheet1[cell_name]
                cell.alignment = Alignment(horizontal='center', vertical='center')

            service_start_count = service_end_count + 1

        server_end_count = server_start_count + server_inc_count - 1
        #"服务器"合并并居中
        column_count = 0
        for i in list(map(chr, range(ord('A'), ord('K') + 1))):
            column_count += 1
            start_cell = str(i) + str(server_start_count)
            end_cell = str(i) + str(server_end_count)
            sheet1.merge_cells(start_cell + ':' + end_cell)
            cell_name = str(i) + str(server_start_count)
            cell = sheet1[cell_name]
            cell.alignment = Alignment(horizontal='center', vertical='center')

        server_start_count = server_end_count + 1

    #写入项目信息
    sheet2['A3'] = project.name
    sheet2['B3'] = project.get_env_display()
    sheet2['C3'] = project.mx_version
    sheet2['A6'] = project.cp_addr
    sheet2['B6'] = project.cp_admin
    sheet2['C6'] = project.cp_password
    sheet2['A9'] = project.monitor_addr
    sheet2['B9'] = project.monitor_admin
    sheet2['C9'] = project.monitor_password
    sheet2['A12'] = project.ops_admin
    sheet2['B12'] = project.ops_password
    sheet2['A15'] = project.remark
    
    #sheet1全表居中
    nrows = sheet1.max_row
    ncols = sheet1.max_column
    for i in range(nrows):
        for j in range(ncols):
            sheet1.cell(row=i+1, column=j+1).alignment = Alignment(horizontal='center', vertical='center')
        
    #生成文件
    file_name = str(project.name) + '-' + str(project.get_env_display()) + '.xlsx'
    book.save('tmp/export/' + file_name)

    file = open('tmp/export/' + file_name, 'rb')
    response = FileResponse(file)
    response['Content-Type'] = 'application/octet-stream'
    response['Content-Disposition'] = 'attachment;filename="{}"'.format(escape_uri_path(file_name))
    return response
